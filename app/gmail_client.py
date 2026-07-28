"""Pull recent email from one or more Gmail accounts via the official API
(read-only scope). Each account has its own OAuth token file — see
scripts/gmail_auth.py and GMAIL_TOKEN_FILES in .env."""
import base64
import logging
import re
from pathlib import Path

from google.auth.transport.requests import Request
from google.oauth2.credentials import Credentials
from googleapiclient.discovery import build

from . import config, db

log = logging.getLogger("gmail")

SCOPES = ["https://www.googleapis.com/auth/gmail.readonly"]


def _credentials(token_file: str) -> Credentials:
    token_path = Path(token_file)
    if not token_path.exists():
        raise RuntimeError(
            f"Gmail token not found at {token_path}. Run scripts/gmail_auth.py "
            "on your own machine for this account, then copy the token file to the server."
        )
    creds = Credentials.from_authorized_user_file(str(token_path), SCOPES)
    if creds.expired and creds.refresh_token:
        creds.refresh(Request())
        token_path.write_text(creds.to_json())
    return creds


def _extract_body(payload: dict) -> str:
    """Walk the MIME tree and return the best-effort plain-text body."""
    if payload.get("mimeType") == "text/plain" and payload.get("body", {}).get("data"):
        return base64.urlsafe_b64decode(payload["body"]["data"]).decode(errors="replace")
    for part in payload.get("parts", []) or []:
        text = _extract_body(part)
        if text:
            return text
    # fall back to text/html stripped very roughly
    if payload.get("mimeType") == "text/html" and payload.get("body", {}).get("data"):
        import re
        html = base64.urlsafe_b64decode(payload["body"]["data"]).decode(errors="replace")
        return re.sub(r"<[^>]+>", " ", html)
    return ""


# ── attachment reading ───────────────────────────────────────────────────────
# Small PDFs / Excels / CSVs get their text extracted and appended to the
# stored mail body, so the AI reads what's INSIDE "PFA" mails (POs,
# checklists, invoices) — previously a blind spot.

MAX_ATTACHMENTS_PER_MAIL = 3
MAX_ATTACHMENT_BYTES = 5 * 1024 * 1024   # 5 MB
MAX_ATTACHMENT_CHARS = 4000              # per attachment, keeps prompts sane

_READABLE_EXT = (".pdf", ".xlsx", ".xlsm", ".csv", ".txt")


def _attachment_text(filename: str, data: bytes) -> str:
    """Best-effort text from one attachment. Empty string if unreadable."""
    name = filename.lower()
    try:
        if name.endswith(".pdf"):
            from io import BytesIO

            from pypdf import PdfReader
            reader = PdfReader(BytesIO(data))
            parts = []
            for page in reader.pages[:8]:
                parts.append(page.extract_text() or "")
            text = "\n".join(parts).strip()
            return text if text else "[scanned PDF — no extractable text]"
        if name.endswith((".xlsx", ".xlsm")):
            from io import BytesIO

            from openpyxl import load_workbook
            wb = load_workbook(BytesIO(data), read_only=True, data_only=True)
            ws = wb.worksheets[0]
            lines = []
            for i, row in enumerate(ws.iter_rows(values_only=True)):
                if i >= 100:
                    break
                cells = [str(c) for c in row if c is not None]
                if cells:
                    lines.append(" | ".join(cells))
            return "\n".join(lines)
        if name.endswith((".csv", ".txt")):
            return data.decode("utf-8", errors="replace")
    except Exception:
        log.exception("could not read attachment %s", filename)
    return ""


def _walk_attachments(payload: dict) -> list:
    """All parts that are real file attachments: (filename, attachmentId, size)."""
    out = []
    for part in payload.get("parts", []) or []:
        fname = part.get("filename") or ""
        body = part.get("body", {}) or {}
        if fname and body.get("attachmentId"):
            out.append((fname, body["attachmentId"], int(body.get("size") or 0)))
        out.extend(_walk_attachments(part))
    return out


def _extract_attachments(service, msg_id: str, payload: dict) -> str:
    """Readable attachments' text, formatted for the AI. '' when none."""
    sections = []
    for fname, att_id, size in _walk_attachments(payload)[:MAX_ATTACHMENTS_PER_MAIL]:
        if not fname.lower().endswith(_READABLE_EXT):
            continue
        if size > MAX_ATTACHMENT_BYTES:
            sections.append(f"[ATTACHMENT: {fname} — too large to read]")
            continue
        try:
            att = service.users().messages().attachments().get(
                userId="me", messageId=msg_id, id=att_id).execute()
            data = base64.urlsafe_b64decode(att.get("data", ""))
        except Exception:
            log.exception("attachment download failed: %s", fname)
            continue
        text = _attachment_text(fname, data)
        if text:
            sections.append(f"[ATTACHMENT: {fname}]\n{text[:MAX_ATTACHMENT_CHARS]}")
    return "\n\n".join(sections)


def _fetch_account(token_file: str, query: str | None = None, max_pages: int = 1) -> int:
    """Fetch mail for a single account. Returns count stored."""
    service = build("gmail", "v1", credentials=_credentials(token_file), cache_discovery=False)
    account = service.users().getProfile(userId="me").execute().get("emailAddress", token_file)
    refs, page_token = [], None
    for _ in range(max_pages):
        resp = service.users().messages().list(
            userId="me", q=query or config.GMAIL_QUERY, maxResults=100,
            pageToken=page_token,
        ).execute()
        refs += resp.get("messages", [])
        page_token = resp.get("nextPageToken")
        if not page_token:
            break
    stored = 0
    for ref in refs:
        msg = service.users().messages().get(
            userId="me", id=ref["id"], format="full"
        ).execute()
        headers = {
            h["name"].lower(): h["value"]
            for h in msg.get("payload", {}).get("headers", [])
        }
        body = _extract_body(msg.get("payload", {}))[:8000]
        att_text = _extract_attachments(service, msg["id"], msg.get("payload", {}))
        if att_text:
            body = (body + "\n\n" + att_text)[:20000]
        sender = headers.get("from", "")
        sender_addr = (re.findall(r"[\w.+-]+@[\w.-]+", sender) or [""])[0].lower()
        if any(pat in sender.lower() for pat in config.IGNORE_SENDERS):
            continue  # automated/noise sender — don't store at all
        direction = (
            "outgoing"
            if "SENT" in msg.get("labelIds", []) or sender_addr in config.OWNER_EMAILS
            else "incoming"
        )
        # Group by the address the mail was originally sent to (so a hub inbox
        # receiving forwards still shows mail under the original business
        # address). Falls back to the connected account.
        to_addr = (re.findall(r"[\w.+-]+@[\w.-]+",
                              headers.get("delivered-to", "") or headers.get("to", ""))
                   or [account])[0].lower()
        if db.save_email(
            gmail_id=msg["id"],
            account=to_addr if direction == "incoming" else account,
            sender=sender,
            subject=headers.get("subject", ""),
            snippet=msg.get("snippet", ""),
            body=body,
            ts=headers.get("date", ""),
            direction=direction,
        ):
            stored += 1
    log.info("gmail[%s]: stored %d new emails", account, stored)
    return stored


def fetch_recent_emails(query: str | None = None, max_pages: int = 1) -> int:
    """Fetch mail across ALL configured accounts. One account failing
    (expired token, network) never blocks the others."""
    total = 0
    for token_file in config.GMAIL_TOKEN_FILES:
        try:
            total += _fetch_account(token_file, query=query, max_pages=max_pages)
        except Exception:
            log.exception("gmail fetch failed for %s — continuing", token_file)
    return total
